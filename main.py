import datetime
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Border, Side,
    Alignment, Font, GradientFill
)
import openpyxl.utils.cell
from openpyxl.chart import BarChart, Reference

from openpyxl import Workbook

"""
базовая "в область" - фазы 1, 6, 7, 8, 9
базовая "в обе стороны" - фазы 2, 10, 11, 12, 13
базовая "по Салтыковской" - фазы 4, 5, 26, 27, 28, 29
"""

# path = r"3412_stages.xlsx"
path_to_save = r"new_3412_stages.xlsx"


# print(wb.sheetnames)
# group1 = ('10', '11', '12', '13')
# group2 = ('26', '27', '28', '29')
# group3 = ('6', '7', '8', '9')
# print(sheet1.max_row)


# Привзяка направлений к фазам
# napr1 = {1, 2, 6, 7, 8, 9, 10, 11, 12, 13}
# napr2 = {2, 10, 11, 12, 13}
# napr3 = {2, 10, 11, 12, 13}
# napr4 = {1, 2, 6, 7, 8, 9, 10, 11, 12, 13}
# napr5 = {1, 6, 7, 8, 9}
# napr6 = {4, 5, 26, 27, 28, 29}
# napr7 = {4, 5, 26, 27, 28, 29}
# napr8 = {4, 5, 26, 27, 28, 29}
# # napr9
# napr10 = {2, 4, 7, 9, 11, 13, 27, 29}
# napr11 = {1, 4, 8, 9, 12, 13, 28, 29}
# napr12 = {2, 4, 7, 9, 11, 13, 27, 29}
# napr13 = {1, 4, 6, 8, 9, 10, 12, 13, 26, 28, 29}
# napr14 = {1, 5, 6, 8, 10, 12, 26, 28}
# napr15 = {1, 5, 6, 8, 10, 12, 26, 28}
# napr16 = {2, 5, 6, 7, 10, 11, 26, 27}
# napr17 = {2, 5, 7, 11, 27}
def read_user_data_from_file():
    with open('report1.txt', encoding='UTF-8') as file:
        napravleniya = dict()
        imenovannye_fazy_napravleniya = dict()
        faza_napravlenie = dict()
        point_new_cyc = []
        headers = ['Типы направлений', 'Точка начала цикла(фаза)', 'Именованные направления',
                   'Принадлежность фаз к именованным направлениям', 'Фаза-направление',
                   'Период-интервал, в минутах']
        flag = False
        for line in file:
            line = line.strip()
            if line in headers or line == 'end':
                flag = line
                continue

            if flag == headers[0] and line != 'end':
                num_napr, name_napr = line.strip().split(': ')
                napravleniya[num_napr] = name_napr
            elif flag == headers[1] and line != 'end':
                # point_new_cyc = list(map(int, line.split(','))) # интовые значения
                point_new_cyc = line.replace(' ', '').split(',')
            elif flag == headers[3]:
                name_napr, num_stage = line.strip().split(': ')
                num_stage = num_stage.replace(' ', '').split(',')
                imenovannye_fazy_napravleniya[name_napr] = num_stage
            elif flag == headers[4]:
                num_stage, num_napr = line.strip().split(': ')
                num_napr = num_napr.replace(' ', '').split(',')
                faza_napravlenie[num_stage] = num_napr

        print(f'napravleniya: {napravleniya}')
        print(f'point_new_cyc: {point_new_cyc}')
        print(f'imenovannye_fazy_napravleniya: {imenovannye_fazy_napravleniya}')
        print(f'faza_napravlenie: {faza_napravlenie}')

class Styles:
    def __init__(self):
        self.color_pink = openpyxl.styles.colors.Color(rgb='00FFCC99')

        self.green_fill1 = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        self.red_fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
        self.green_fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
        self.ligh_pink = PatternFill(fgColor=self.color_pink)

    def boarder_all_side(self):
        thin = Side(border_style="thin")
        return Border(top=thin, left=thin, right=thin, bottom=thin)

    def fill_red(self):
        pass




class ReadUserData:
    def __init__(self, filename='report1.txt'):
        self.filename = filename
        self.path_to_excel = None
        self.kolichestvo_napravleniy = None
        self.napravleniya = dict()
        self.imenovannye_fazy_napravleniya = dict()
        self.faza_napravlenie = dict()
        self.point_new_cyc = []
        self.report_interval = 15
        self.read_data()

    def read_data(self):
        with open(self.filename, encoding='UTF-8') as file:
            headers = ['Путь к excel файлу', 'Типы направлений', 'Точка начала цикла(фаза)', 'Именованные направления',
                       'Принадлежность именованных направлений к фазам', 'Фаза-направление',
                       'Период-интервал, в минутах']
            flag = False
            for line in file:
                line = line.strip()
                if line in headers or line == 'end':
                    flag = line
                    continue

                if flag == headers[0] and line != 'end':
                    self.path_to_excel = line
                elif flag == headers[1] and line != 'end':
                    num_napr, name_napr = line.strip().split(': ')
                    self.napravleniya[num_napr] = name_napr
                elif flag == headers[2] and line != 'end':
                    # point_new_cyc = list(map(int, line.split(','))) # интовые значения
                    self.point_new_cyc = line.replace(' ', '').split(',')
                elif flag == headers[4]:
                    name_napr, num_stage = line.strip().split(': ')
                    num_stage = num_stage.replace(' ', '').split(',')
                    self.imenovannye_fazy_napravleniya[name_napr] = num_stage
                elif flag == headers[5]:
                    num_stage, num_napr = line.strip().split(': ')
                    num_napr = num_napr.replace(' ', '').split(',')
                    self.faza_napravlenie[num_stage] = num_napr
                elif flag == headers[6]:
                    self.report_interval = line

            print(f'self.path_to_excel: {self.path_to_excel}')
            print(f'napravleniya: {self.napravleniya}')
            print(f'point_new_cyc: {self.point_new_cyc}')
            print(f'imenovannye_fazy_napravleniya: {self.imenovannye_fazy_napravleniya}')
            print(f'faza_napravlenie: {self.faza_napravlenie}')
            print(f'self.report_interval: {self.report_interval}')

        self.kolichestvo_napravleniy = len(self.napravleniya)
        print(f'self.kolichestvo_napravleniy: {self.kolichestvo_napravleniy}')


class TableConstructorFirstSheet:
    def __init__(self, groups, start_column=6):
        self.groups = groups

        self.napr_name_cell = dict()
        self.naimenovanie_napravleniy = dict()

        self.time_column = 1
        self.stage_column = 2
        self.dlitelnost_column = 3
        self.num_cyc_column = 4

        self.start_column = self.num_cyc_column + user_data.kolichestvo_napravleniy + 1

        self.mode_column = self.start_column + 1
        self.imenovannoe_napravlenie_column = self.start_column + 2
        self.napr_v_faze_column = self.start_column + 3
        # self.num_cyc_column = start_column + 4
        self.common_cyc_column = self.start_column + 5  # это как проверка счётчика циклов
        self.cyc_interval_column = self.start_column + 6

    def make_table_header(self, sheet):
        """ Метод формирует заголовки(имена ячеек) таблицы 1 строки"""

        # Формируем имена колонок направлений вида Т1(1 Траспортное)/П2(2 Пешеходное) и т.д.
        column = self.num_cyc_column + 1
        for k, v in user_data.napravleniya.items():
            sheet.cell(row=1, column=column).value = f'{v[0]}{k}'
            name_cell = f'{openpyxl.utils.cell.get_column_letter(column)}'
            self.napr_name_cell[k] = name_cell
            column += 1
        print(f'self.napr_name_cell: {self.napr_name_cell}')

        # Формируем названия колонок после направлений
        sheet.cell(row=1, column=self.start_column).value = 'Резерв'
        sheet.cell(row=1, column=self.mode_column).value = 'Режим'
        sheet.cell(row=1, column=self.imenovannoe_napravlenie_column).value = 'Название направления'
        sheet.cell(row=1, column=self.napr_v_faze_column).value = 'Номера направлений в фазе'
        sheet.cell(row=1, column=self.num_cyc_column).value = 'Номер цикла'
        sheet.cell(row=1, column=self.common_cyc_column).value = 'Общее кол-во циклов'
        sheet.cell(row=1, column=self.cyc_interval_column).value = 'Кол-во циклов, за час'

        offset = self.cyc_interval_column + 1
        for k in self.groups.keys():
            self.naimenovanie_napravleniy[k] = f'{openpyxl.utils.cell.get_column_letter(offset)}'
            sheet.cell(row=1, column=offset).value = k
            offset += 1
        print(f'self.naimenovanie_napravleniy: {self.naimenovanie_napravleniy}')


class TableConstructorSecondSheet:
    def __init__(self):
        # self.sheet2 = wb['Отчёт1']

        self.time_column = 1
        self.cyc_common_column = self.time_column + 1
        self.cyc_interval_column = self.cyc_common_column + 1
        self.naimenovanie_napravleniy = dict()



    def make_table_header(self, sheet):

        sheet.cell(row=1, column=self.time_column).value = 'Время/Период'
        sheet.cell(row=1, column=self.cyc_common_column).value = 'Номер цикла(общий счёт)'
        sheet.cell(row=1, column=self.cyc_interval_column).value = f'Кол-во циклов за {user_data.report_interval} мин'

        column = self.cyc_interval_column + 1
        for name_napr in user_data.imenovannye_fazy_napravleniya:
            sheet.cell(row=1, column=column).value = name_napr
            name_cell = f'{openpyxl.utils.cell.get_column_letter(column)}'
            self.naimenovanie_napravleniy[name_napr] = name_cell
            column += 1
        print(f'self.naimenovanie_napravleniy из TableConstructorSecondSheet: {self.naimenovanie_napravleniy}')


class MakeReport:
    def __init__(self, wb):
        self.sheet1 = wb['Page1']
        self.sheet2 = wb['Отчёт1']
        self.sheet3 = wb['Отчёт2']

        self.curr_row = 2

        self.cnt_all_cyc = 1
        self.cnt_cyc_interval = 1

        self.curr_stage = 0
        self.prev_stage = 0

        self.curr_hour = 0
        self.prev_hour = 0

        self.minutes_cnt = 0

        self.interval_point = False

        self.cnt_common_time = 0
        self.time_kazhdoe_napravlenie = {k: 0 for k in user_data.napravleniya}
        print(f'time_kazhdoe_napravlenie: {self.time_kazhdoe_napravlenie}')
        self.time_cnt = {k: 0 for k in user_data.imenovannye_fazy_napravleniya}
        print(f'self.time_cnt: {self.time_cnt}')

        self.point_new_cyc = user_data.point_new_cyc
        self.time_interval = int(user_data.report_interval)
        self.all_intervals_data = []

    def processing_excel_file(self):
        sheet = self.sheet1

        time_column = table_header_sheet1.time_column
        num_stage_column = table_header_sheet1.stage_column
        dlitelnost_column = table_header_sheet1.dlitelnost_column

        for curr_row in range(2, self.sheet1.max_row + 1):
            if curr_row > 2:
                self.curr_row = curr_row

                self.curr_stage = sheet.cell(row=self.curr_row, column=num_stage_column).value
                self.prev_stage = sheet.cell(row=self.curr_row - 1, column=num_stage_column).value

                self.curr_hour = sheet.cell(row=self.curr_row, column=time_column).value.time().hour
                self.prev_hour = sheet.cell(row=self.curr_row - 1, column=time_column).value.time().hour

                self.curr_minute = int(sheet.cell(row=self.curr_row, column=time_column).value.time().minute)
                self.prev_minute = int(sheet.cell(row=self.curr_row - 1, column=time_column).value.time().minute)

                self.curr_dlitelnost_stage = sheet.cell(row=self.curr_row, column=dlitelnost_column).value
                # print(f'self.curr_minute: {self.curr_minute}')
                # print(f'self.prev_minute: {self.prev_minute}')

                # curr_cell_val_hour = sheet1.cell(row=row, column=column_datetime).value.time().hour
                # prev_cell_val_hour = sheet1.cell(row=row - 1, column=column_datetime).value.time().hour

                self.calc_all_cyc()
                self.write_num_curr_cyc()

                name, stages = self.read_stages_and_names()
                self.write_num_and_name_napravleniy(name, stages)
                self.filling_the_cell()

                self.sum_time()
                # self.minutes_count()
                if self.check_interval_point():
                    self.write_interval_data()
                    # self.write_to_sheet2()

                    self.reset_val()
                    # print(f'self.time_cnt : {self.time_cnt}')

            else:
                self.write_data_in_1_row()

    def calc_all_cyc(self):
        """ Метод подсчёта текущего номера цикла """
        if self.curr_stage in self.point_new_cyc and self.prev_stage not in self.point_new_cyc:
            self.cnt_all_cyc += 1
            self.cnt_cyc_interval += 1
        elif self.curr_stage in ('-1', '-2'):
            self.cnt_all_cyc += 1
            self.cnt_cyc_interval += 1


    def read_stages_and_names(self):
        """ Метод вычисляет какие номера фаз привязаны к именам направления движения """
        for k, val in user_data.imenovannye_fazy_napravleniya.items():
            if self.curr_stage in val:
                napravleniya = user_data.faza_napravlenie.get(self.curr_stage)
                return k, ','.join(napravleniya)
        return "NaN", 'NaN'

        # Старая версия с дублями записей(если подряд одинаковые)
        # for k, val in user_data.imenovannye_fazy_napravleniya.items():
        #     if self.curr_stage in val:
        #         napravleniya = user_data.faza_napravlenie.get(self.curr_stage)
        #         return k, ','.join(napravleniya)
        # return "NaN", 'NaN'

    def write_num_curr_cyc(self):
        """ Метод записывает в ячейку номер текущего цикла """
        self.sheet1.cell(row=self.curr_row, column=table_header_sheet1.num_cyc_column).value = self.cnt_all_cyc

    def write_data_in_1_row(self):
        self.write_num_curr_cyc()

    def write_num_and_name_napravleniy(self, name, stages):
        prev_name = self.sheet1.cell(row=self.curr_row - 1, column=table_header_sheet1.imenovannoe_napravlenie_column).value

        if prev_name != name:
            self.sheet1.cell(row=self.curr_row, column=table_header_sheet1.imenovannoe_napravlenie_column).value = name
            self.sheet1.cell(row=self.curr_row, column=table_header_sheet1.napr_v_faze_column).value = stages

        # Старая версия с дублями записей(если подряд одинаковые)
        # self.sheet.cell(row=self.curr_row, column=table_header.imenovannoe_napravlenie_column).value = name
        # self.sheet.cell(row=self.curr_row, column=table_header.napr_v_faze_column).value = stages

    def reset_val(self):
        self.cnt_cyc_interval = 0
        # Сброс общего времени для именованного направления
        for name in self.time_cnt:
            self.time_cnt[name] = 0
        # Сброс общего времени для направления
        for num_napr in self.time_kazhdoe_napravlenie:
            self.time_kazhdoe_napravlenie[num_napr] = 0

    def check_interval_point(self):

        res = self.sheet1.cell(row=self.curr_row - 1, column=6).value
        if self.time_interval == 60:
            if self.curr_hour != self.prev_hour:
                self.interval_point = True
                return True


        elif self.time_interval < 60 and self.curr_minute > 0:
            if self.curr_minute % self.time_interval == 0 and res != self.time_interval:
                self.sheet1.cell(row=self.curr_row, column=table_header_sheet1.start_column).value = f'Mod {self.time_interval}'
        elif 60 < self.time_interval < 120:
            pass

        # print(f'res: {res}')
        # if self.curr_minute > 0:
        #     print(f'self.time_interval: {self.time_interval}')
        #     print(f'self.curr_minute: {self.curr_minute}')
        #
        #
        #     print(f'self.time_interval % self.curr_minute: {self.time_interval % self.curr_minute}')

    def minutes_count(self):
        if self.curr_minute > self.prev_minute:
            self.minutes_cnt += 1
        elif self.curr_minute == 0 and self.prev_minute == 59:
            self.minutes_cnt += 1

    def filling_the_cell(self):
        """ Заливка зелёным/красным цветом ячейки направления """

        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        for k, v in table_header_sheet1.napr_name_cell.items():
            name_cell = f'{v}{self.curr_row}'
            if self.curr_stage != '-1' and k is not None and k in user_data.faza_napravlenie.get(str(self.curr_stage)):
                self.sheet1[name_cell].fill = green_fill
                self.time_kazhdoe_napravlenie[k] += self.curr_dlitelnost_stage
            elif self.curr_stage not in ('-1', '-2'):
                self.sheet1[name_cell].fill = red_fill

    def sum_time(self):
        for name_napr, stage in user_data.imenovannye_fazy_napravleniya.items():
            if self.curr_stage in stage:
                self.time_cnt[name_napr] += self.curr_dlitelnost_stage
                return

    def write_interval_data(self):

        style = Styles()

        # Запись в лист 1
        self.sheet1.cell(row=self.curr_row - 1, column=table_header_sheet1.common_cyc_column).value = self.cnt_all_cyc
        self.sheet1.cell(row=self.curr_row - 1, column=table_header_sheet1.cyc_interval_column).value = self.cnt_cyc_interval
        # Запись в лист 2
        time = self.sheet1.cell(row=self.curr_row, column=table_header_sheet1.time_column).value
        row_for_sheet2 = self.sheet2.max_row + 1
        self.sheet2.cell(row=row_for_sheet2, column=table_header_sheet2.time_column).value = time
        self.sheet2[f'A{row_for_sheet2}'].fill = style.green_fill1
        self.sheet2.cell(row=row_for_sheet2, column=table_header_sheet2.cyc_common_column).value = self.cnt_all_cyc
        self.sheet2.cell(row=row_for_sheet2,
                         column=table_header_sheet2.cyc_interval_column).value = self.cnt_cyc_interval

        for name, sum_time in self.time_cnt.items():
            cell_total = f'{table_header_sheet1.naimenovanie_napravleniy.get(name)}{self.curr_row - 1}'
            average_time = sum_time // self.cnt_cyc_interval
            self.sheet1[cell_total] = average_time
            cell_average = f'{table_header_sheet1.naimenovanie_napravleniy.get(name)}{self.curr_row - 2}'
            self.sheet1[cell_average] = sum_time
            # Запись в лист 2
            cell_total2 = f'{table_header_sheet2.naimenovanie_napravleniy.get(name)}{row_for_sheet2}'
            self.sheet2[cell_total2] = sum_time
            cell_average2 = f'{table_header_sheet2.naimenovanie_napravleniy.get(name)}{row_for_sheet2 + 1}'
            self.sheet2[cell_average2] = average_time

        # Запись в лист 2 таблицы для каждого направления
        row_for_sheet2 = self.sheet2.max_row

        # Создание гистограммы
        # create data for plotting
        values = Reference(self.sheet2, min_col=4, min_row=row_for_sheet2,
                           max_col=3 + len(self.time_cnt), max_row=row_for_sheet2)

        # Create object of BarChart class
        chart = BarChart()

        # adding data to the Bar chart object
        chart.add_data(values)

        # set the title of the chart
        chart.title = " BAR-CHART "

        # set the title of the x-axis
        chart.x_axis.title = " X_AXIS "

        # set the title of the y-axis
        chart.y_axis.title = " Y_AXIS "


        chart.width = len(self.time_cnt) * 2
        chart.height = user_data.kolichestvo_napravleniy // 2
        # add chart to the sheet
        # the top-left corner of a chart
        # is anchored to cell E2 .
        self.sheet2.add_chart(chart, f"I{row_for_sheet2 + 2}")




        # Запись в лист 2 таблицы для каждого направления
        row_for_sheet2 = self.sheet2.max_row + 2
        self.sheet2[f'A{row_for_sheet2}'] = 'Номер'
        self.sheet2[f'A{row_for_sheet2}'].border = style.boarder_all_side()
        self.sheet2[f'B{row_for_sheet2}'] = 'Тип'
        self.sheet2[f'B{row_for_sheet2}'].border = style.boarder_all_side()
        self.sheet2[f'C{row_for_sheet2}'] = f'Время всего за {user_data.report_interval} мин'
        self.sheet2[f'C{row_for_sheet2}'].border = style.boarder_all_side()
        self.sheet2[f'D{row_for_sheet2}'] = f'Среднее время в цикле, за {user_data.report_interval} мин'
        self.sheet2[f'D{row_for_sheet2}'].border = style.boarder_all_side()
        for num_napr, time in self.time_kazhdoe_napravlenie.items():
            row_for_sheet2 += 1
            self.sheet2[f'A{row_for_sheet2}'] = num_napr
            self.sheet2[f'A{row_for_sheet2}'].border = style.boarder_all_side()


            self.sheet2[f'B{row_for_sheet2}'] = user_data.napravleniya.get(num_napr)
            self.sheet2[f'B{row_for_sheet2}'].border = style.boarder_all_side()
            self.sheet2[f'C{row_for_sheet2}'] = time
            self.sheet2[f'C{row_for_sheet2}'].border = style.boarder_all_side()
            self.sheet2[f'D{row_for_sheet2}'] = time // self.cnt_cyc_interval
            self.sheet2[f'D{row_for_sheet2}'].border = style.boarder_all_side()






    def write_to_sheet2(self):
        pass





user_data = ReadUserData()
wb = openpyxl.load_workbook(user_data.path_to_excel)
wb.create_sheet('Отчёт1')
wb.create_sheet('Отчёт2')

sheet1 = wb['Page1']
sheet2 = wb['Отчёт1']
sheet3 = wb['Отчёт2']

table_header_sheet1 = TableConstructorFirstSheet(user_data.imenovannye_fazy_napravleniya, )
table_header_sheet1.make_table_header(sheet1)

table_header_sheet2 = TableConstructorSecondSheet()
table_header_sheet2.make_table_header(sheet2)

report = MakeReport(wb)
report.processing_excel_file()

wb.save(path_to_save)
sys.exit()

# sheet1['I1'] = 'Общее кол-во циклов'
# sheet1['J1'] = 'Кол-во циклов, за час'
# sheet1['L1'] = 'В область'
# sheet1['M1'] = 'В обе стороны'
# sheet1['N1'] = 'По салтыковке'

print(f'user_data ---> {user_data.imenovannye_fazy_napravleniya}')

vse_napravleniya_time = [0 if i > 0 else 'Длительности направлений' for i in range(18)]


def write_dlitelnost_napr(napravleniya, time):
    for i in napravleniya:
        vse_napravleniya_time[i] += time


def write_table(start_row):
    print(f'start_row: {start_row}')

    row = start_row - 20
    for num_napr in range(18):
        if num_napr > 0:
            sheet1.cell(row=row, column=column_cnt_cyc_hour).value = num_napr
            sheet1.cell(row=row, column=column_cnt_cyc_hour + 1).value = table_napravl[num_napr]
            sheet1.cell(row=row, column=column_cnt_cyc_hour + 2).value = vse_napravleniya_time[num_napr]
        row += 1


# Таблица направлений
table_napr = {1: 'Транспортное',
              2: 'Транспортное',
              3: 'Пешеходное',
              4: 'Пешеходное',
              5: 'Поворотное',
              6: 'Пешеходное',
              7: 'Транспортное',
              8: 'Транспортное',
              10: 'Поворотное',
              11: 'Поворотное',
              12: 'Поворотное',
              13: 'Поворотное',
              14: 'Пешеходное',
              15: 'Пешеходное',
              16: 'Пешеходное',
              17: 'Пешеходное',
              }
table_napravl = ['Название направлений',
                 'Транспортное',
                 'Транспортное',
                 'Пешеходное',
                 'Пешеходное',
                 'Поворотное',
                 'Пешеходное',
                 'Транспортное',
                 'Транспортное',
                 '-----',
                 'Поворотное',
                 'Поворотное',
                 'Поворотное',
                 'Поворотное',
                 'Пешеходное',
                 'Пешеходное',
                 'Пешеходное',
                 'Пешеходное',
                 ]

print(len(table_napravl))
print(len(vse_napravleniya_time))

# Привязка фаз к наименованию группы(в обе стороны, по салтыковке, в область)
vse_v_oblast = {'1', '6', '7', '8', '9'}
vse_v_obe_storony = {'2', '10', '11', '12', '13'}
vse_saltykovka = {'4', '5', '26', '27', '28', '29'}

va_v_oblast = {'6', '7', '8', '9'}
va_v_obe_storony = {'10', '11', '12', '13'}
va_saltykovka = {'26', '27', '28', '29'}

ft_v_oblast = {'1'}
ft_v_obe_storony = {'2'}
ft_saltykovka = {'4', '5'}

va_stages = va_v_oblast | va_v_obe_storony | va_saltykovka
# print(f'va_stages: {va_stages}')
ft_stages = ft_v_oblast | ft_v_obe_storony | ft_saltykovka
# print(f'ft_stages: {ft_stages}')
stages_for_change_cnt_cyc = vse_v_oblast | vse_saltykovka

column_datetime = 1
column_stage = 2
column_dlitelnost = 3
column_cyc_counter = 5
column_mode = 6
column_napravlenie = 7
column_end_hour = 8
column_cnt_cyc_hour = 9
column_cnt_cyc_hour_supervisor = 10
column_common_time_v_oblast = 12
column_common_time_v_obe_storony = 13
column_common_time_saltykovka = 14

sheet1['I1'] = 'Общее кол-во циклов'
sheet1['J1'] = 'Кол-во циклов, за час'
sheet1['L1'] = 'В область'
sheet1['M1'] = 'В обе стороны'
sheet1['N1'] = 'По салтыковке'

sheet1.cell(row=2, column=column_cyc_counter).value = 1
sheet1.cell(row=2, column=column_mode).value = 'адаптива'
sheet1.cell(row=2, column=column_napravlenie).value = 'в обе стороны'

cnt_cyc = 1
curr_cnt_cyc = 1
cnt_cyc_hour_supervisor = 0
flag_va = True
flag_allow_to_change_cnt_cyc = True

common_time_v_oblast = 0
common_time_v_obe_storony = 0
common_time_saltykovka = 0

# print(type(sheet1.cell(row=2, column=column_datetime).value))
# print(sheet1.cell(row=2, column=column_datetime).value.time().hour)

for row in range(3, sheet1.max_row + 1):
    stage_val_curr = sheet1.cell(row=row, column=column_stage).value
    stage_val_prev = sheet1.cell(row=row - 1, column=column_stage).value

    curr_cell_val_hour = sheet1.cell(row=row, column=column_datetime).value.time().hour
    prev_cell_val_hour = sheet1.cell(row=row - 1, column=column_datetime).value.time().hour
    # stage_val_prev = sheet1.cell(row=row-1, column=column_stage).value
    # stage_val_prev_prev = sheet1.cell(row=row-2, column=column_stage).value

    curr_dlitelnost_val = sheet1.cell(row=row, column=column_dlitelnost).value

    # Проверка режима va/ft
    if stage_val_curr in va_stages:
        sheet1.cell(row=row, column=column_mode).value = 'адаптива'
        flag_va = True
    elif stage_val_curr in ft_stages:
        sheet1.cell(row=row, column=column_mode).value = 'фикс'
        flag_va = False

    # Проверка, можем ли менять счётчик фаз(была ли фаза из объединения множеств vse_v_oblast | vse_saltykovka)

    # if not flag_allow_to_change_cnt_cyc and stage_val_curr in stages_for_change_cnt_cyc:
    #   flag_allow_to_change_cnt_cyc = True

    # Проверяем условие, когда можно при нахождении основной фазы прибавить 1 к счётчику цикла
    if not flag_allow_to_change_cnt_cyc and stage_val_curr in stages_for_change_cnt_cyc:
        flag_allow_to_change_cnt_cyc = True
    # Если значение фазы = -1 или -2, прибавляет с счётчику циклов 1
    if stage_val_curr in ('-1', '-2'):
        cnt_cyc += 1
        curr_cnt_cyc += 1

    # Считаем и записываем номер цикла
    if flag_va:
        if flag_allow_to_change_cnt_cyc and stage_val_curr in va_v_obe_storony:
            cnt_cyc += 1
            curr_cnt_cyc += 1
            flag_allow_to_change_cnt_cyc = False
        sheet1.cell(row=row, column=column_cyc_counter).value = cnt_cyc
    else:
        if flag_allow_to_change_cnt_cyc and stage_val_curr in ft_v_oblast:
            cnt_cyc += 1
            curr_cnt_cyc += 1
            flag_allow_to_change_cnt_cyc = False
        sheet1.cell(row=row, column=column_cyc_counter).value = cnt_cyc

    # Записываем в колонку 7(G) тип направления
    # if stage_val_curr in vse_v_obe_storony and stage_val_prev not in vse_v_obe_storony:
    #     sheet1.cell(row=row, column=column_napravlenie).value = 'в обе стороны'
    # elif stage_val_curr in vse_saltykovka and stage_val_prev not in vse_saltykovka:
    #     sheet1.cell(row=row, column=column_napravlenie).value = 'по салтыковке'
    # elif stage_val_curr in vse_v_oblast and stage_val_prev not in vse_v_oblast:
    #     sheet1.cell(row=row, column=column_napravlenie).value = 'в область'
    if stage_val_curr in vse_v_obe_storony:
        sheet1.cell(row=row, column=column_napravlenie).value = 'в обе стороны'
    elif stage_val_curr in vse_saltykovka:
        sheet1.cell(row=row, column=column_napravlenie).value = 'по салтыковке'
    elif stage_val_curr in vse_v_oblast:
        sheet1.cell(row=row, column=column_napravlenie).value = 'в область'

    # Счёт времени
    if curr_cell_val_hour > prev_cell_val_hour:
        sheet1.cell(row=row - 1, column=column_end_hour).value = 'тест'
        sheet1.cell(row=row - 1, column=column_cnt_cyc_hour).value = curr_cnt_cyc
        cnt_cyc_hour_supervisor += curr_cnt_cyc
        sheet1.cell(row=row - 1, column=column_cnt_cyc_hour_supervisor).value = cnt_cyc_hour_supervisor

        sheet1.cell(row=row - 1, column=column_common_time_v_oblast).value = common_time_v_oblast
        sheet1.cell(row=row - 1, column=column_common_time_v_obe_storony).value = common_time_v_obe_storony
        sheet1.cell(row=row - 1, column=column_common_time_saltykovka).value = common_time_saltykovka

        # Расчёт среднего времени
        average_time_v_oblast = common_time_v_oblast // curr_cnt_cyc
        average_time_v_obe_storony = common_time_v_obe_storony // curr_cnt_cyc
        average_time_saltykovka = common_time_saltykovka // curr_cnt_cyc

        sheet1.cell(row=row - 2, column=column_common_time_v_oblast).value = average_time_v_oblast
        sheet1.cell(row=row - 2, column=column_common_time_v_obe_storony).value = average_time_v_obe_storony
        sheet1.cell(row=row - 2, column=column_common_time_saltykovka).value = average_time_saltykovka

        # Время для каждого направления

        write_table(start_row=row)

        curr_cnt_cyc = 0
        common_time_v_oblast = 0
        common_time_v_obe_storony = 0
        common_time_saltykovka = 0

        for i in range(1, len(vse_napravleniya_time)):
            vse_napravleniya_time[i] = 0

    elif row == sheet1.max_row:
        sheet1.cell(row=row, column=column_end_hour).value = 'тест'
        sheet1.cell(row=row, column=column_cnt_cyc_hour).value = curr_cnt_cyc
        cnt_cyc_hour_supervisor += curr_cnt_cyc
        sheet1.cell(row=row, column=column_cnt_cyc_hour_supervisor).value = cnt_cyc_hour_supervisor

        sheet1.cell(row=row, column=column_common_time_v_oblast).value = common_time_v_oblast
        sheet1.cell(row=row, column=column_common_time_v_obe_storony).value = common_time_v_obe_storony
        sheet1.cell(row=row, column=column_common_time_saltykovka).value = common_time_saltykovka

        # Расчёт среднего времени
        average_time_v_oblast = common_time_v_oblast // curr_cnt_cyc
        average_time_v_obe_storony = common_time_v_obe_storony // curr_cnt_cyc
        average_time_saltykovka = common_time_saltykovka // curr_cnt_cyc

        sheet1.cell(row=row - 1, column=column_common_time_v_oblast).value = average_time_v_oblast
        sheet1.cell(row=row - 1, column=column_common_time_v_obe_storony).value = average_time_v_obe_storony
        sheet1.cell(row=row - 1, column=column_common_time_saltykovka).value = average_time_saltykovka

        # Обнуление данных за час
        curr_cnt_cyc = 0
        common_time_v_oblast = 0
        common_time_v_obe_storony = 0
        common_time_saltykovka = 0

        for i in range(1, len(vse_napravleniya_time)):
            vse_napravleniya_time[i] = 0

    # sheet1.cell(row=sheet1.max_row, column=column_cnt_cyc_hour_supervisor).value = cnt_cyc_hour_supervisor

    # Среднее значение времени для каждого из трех типов направлений как общее время для направления за час,
    # деленое на количество циклов
    if stage_val_curr in vse_v_obe_storony:
        common_time_v_obe_storony += sheet1.cell(row=row, column=column_dlitelnost).value
    elif stage_val_curr in vse_saltykovka:
        common_time_saltykovka += sheet1.cell(row=row, column=column_dlitelnost).value
    elif stage_val_curr in vse_v_oblast:
        # print(f'{sheet1.cell(row=row, column=column_dlitelnost).value}')
        common_time_v_oblast += sheet1.cell(row=row, column=column_dlitelnost).value

    # Время для направлений:
    stage_val_curr_int = int(stage_val_curr)
    if stage_val_curr_int == 1:
        write_dlitelnost_napr([1, 4, 5, 11, 13, 14, 15], curr_dlitelnost_val)

    elif stage_val_curr_int == 2:
        write_dlitelnost_napr([1, 2, 6, 7, 8, 9, 10, 11, 12, 13], curr_dlitelnost_val)

    elif stage_val_curr_int == 4:
        write_dlitelnost_napr([6, 7, 8, 10, 11, 12, 13], curr_dlitelnost_val)

    elif stage_val_curr_int == 5:
        write_dlitelnost_napr([6, 7, 8, 14, 15, 16, 17], curr_dlitelnost_val)

    elif stage_val_curr_int == 6:
        write_dlitelnost_napr([1, 5, 4, 13, 14, 15, 16], curr_dlitelnost_val)


    elif stage_val_curr_int == 7:
        write_dlitelnost_napr([1, 5, 4, 10, 12, 16, 17], curr_dlitelnost_val)


    elif stage_val_curr_int == 8:
        write_dlitelnost_napr([1, 5, 4, 11, 13, 14, 15], curr_dlitelnost_val)

    elif stage_val_curr_int == 9:
        write_dlitelnost_napr([1, 5, 4, 10, 11, 12, 13], curr_dlitelnost_val)


    elif stage_val_curr_int == 10:
        write_dlitelnost_napr([1, 2, 3, 4, 13, 14, 15, 16], curr_dlitelnost_val)


    elif stage_val_curr_int == 11:
        write_dlitelnost_napr([1, 2, 3, 4, 10, 12, 16, 17], curr_dlitelnost_val)


    elif stage_val_curr_int == 12:
        write_dlitelnost_napr([1, 2, 3, 4, 11, 13, 14, 15], curr_dlitelnost_val)


    elif stage_val_curr_int == 13:
        write_dlitelnost_napr([1, 2, 3, 4, 10, 11, 12, 13], curr_dlitelnost_val)


    elif stage_val_curr_int == 26:
        write_dlitelnost_napr([6, 7, 8, 13, 14, 15, 16], curr_dlitelnost_val)


    elif stage_val_curr_int == 27:
        write_dlitelnost_napr([6, 7, 8, 10, 12, 16, 17], curr_dlitelnost_val)

    elif stage_val_curr_int == 28:
        write_dlitelnost_napr([6, 7, 8, 10, 12, 16, 17], curr_dlitelnost_val)

    elif stage_val_curr_int == 29:
        write_dlitelnost_napr([6, 7, 8, 10, 11, 12, 13], curr_dlitelnost_val)

    # print(f'row={row}, val={sheet1.cell(row=row, column=2).value}')
    # print(f'type = {type(sheet1.cell(row=row, column=2).value)}')
    # if row % 2 == 0:
    #     sheet1.cell(row=row, column=5).value = 'Чёт'
    # else:
    #     sheet1.cell(row=row, column=5).value = 'Нечет'

wb.save(path_to_save)
# sheet1.save(path)
